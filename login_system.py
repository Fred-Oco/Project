from openpyxl import load_workbook

wb = load_workbook(filename='Login_Data.xlsx')  # loading excel for date storage
environment = wb['工作表1']
username_row = environment['B']
password_row = environment['C']

def Get_Account_Number():
    number = environment['A']
    result = number[len(number) - 1]
    return result.value

Acc_count = Get_Account_Number()

def Get_User_Input():
    User = Account(input('Login or Create A New Account? ').strip().title(), input('Username: '), input('Password: '))  # asking for user information
    User.Action()

class Account:
    def __init__(self,option, username, password):
        self.option = option
        self.username = username
        self.password = password

    def Action(self):
        if self.option == 'Login':
            check = self.login()
            if check is True:
                print('Logged in')
            else:
                print('Try again')
                Get_User_Input()
        elif self.option == 'Create':
            if self.Check_Duplicate() is True:
                Acc_count = Get_Account_Number()
                count = environment.cell(Acc_count + 2, 1)
                N_username_cell = environment.cell(Acc_count + 2, 2)
                N_password_cell = environment.cell(Acc_count + 2, 3)
                N_username_cell.value = self.username
                N_password_cell.value = self.password
                count.value = Acc_count + 1
                wb.save(filename="Login_Data.xlsx")
                print('Success')
            else:
                print('Same Username Detected')
            Get_User_Input()
        elif self.option == 'Exit':
            print('Closing...')
            quit()
        else:
            print('Wrong Input')
            Get_User_Input()

    def Check_Duplicate(self):  # check for duplicate username
        Acc_count = Get_Account_Number()
        for username in range(Acc_count + 1):
            if username_row[username].value == self.username:
                return False
        return True

    def login(self):
        Acc_count = Get_Account_Number()
        if Acc_count == 1:
            if username_row[1].value == self.username and password_row[1].value == self.password:
                return True
            else:
                return False
        else:
            for acc in range(1, Acc_count):

                if username_row[acc].value == self.username and password_row[acc].value == self.password:
                    return True
                else:
                    return False


Get_User_Input()



